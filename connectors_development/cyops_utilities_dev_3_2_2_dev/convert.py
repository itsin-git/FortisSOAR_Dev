"""
Steps related to convert data from one format to another.
"""
import ast
import json
import re
import os
import xmltodict
import xlrd
import jsonpatch
import csv
import markdown
from collections import OrderedDict
from bs4 import BeautifulSoup
from html import unescape
from .errors.error_constants import *
from django.conf import settings
from django.core.serializers.json import DjangoJSONEncoder
from openpyxl import load_workbook

from connectors.core.connector import get_logger, ConnectorError
from .files import check_file_traversal, download_file_from_cyops

logger = get_logger("cyops_utilities.builtins.convert")


def map_json(json_data, patch, *args, **kwargs):
    """
    Applies a `JSON patch <https://tools.ietf.org/html/rfc6902>`_ to the given
    json

   :param str json_data: the json to transform
   :param str patch: the json for transformation

   :return: the newly formed json
   :rtype: dict
    """
    return _convert_json(json_data, patch)


def _convert_json(json_data, patch):
    """
    Converts json as per RFC 6902

   :param str json_data: the json to transform
   :param str patch: the json for transformation

   :return: the newly formed json
   :rtype: dict
    """
    # omg the error noise
    if type(json_data) == str:
        try:
            obj = json.loads(json_data, strict=False)
        except ValueError:
            try:
                obj = ast.literal_eval(json_data)
            except ValueError as e:
                logger.error('{0} ERROR :: {1}'.format(cs_connector_utility_6,str(e)))
                raise ConnectorError(cs_connector_utility_6)
    elif type(json_data) == dict:
        obj = json_data
    else:
        logger.error(cs_connector_utility_6)
        raise ConnectorError(cs_connector_utility_6)

    try:
        result = jsonpatch.apply_patch(obj, patch)
    except ValueError:
        logger.error(cs_connector_utility_6)
        raise ConnectorError(cs_connector_utility_6)

    return result


def xml_to_dictionary(xml=None, *args, **kwargs):
    """
    Converts an XML string into a dictionary

   :param str xml: The xml to be converted into dictionary
   :return: The converted dictionary
   :rtype: dict
    """
    if not xml:
        raise ConnectorError("%s"%cs_connector_utility_1.format('XML'))
    clean_xml = re.sub('\s+', ' ', xml).strip()
    return json.loads(json.dumps(xmltodict.parse(clean_xml)))


def convert_to_json(file_type=None, file_name=None, iri=None, *args, **kwargs):
    """
    Converts an XML string into a dictionary

   :param str file_type: Type of the file i.e XML or CSV
   :param str file_name: The xml filename in /tmp directory
   :return: The converted dictionary
   :rtype: dict
    """
    if not file_name and iri:
        metadata = download_file_from_cyops(iri, None, *args, **kwargs)
        file_name = metadata.get('cyops_file_path')
    if file_name:
        check_file_traversal(file_name)
        abs_filename = os.path.join(settings.TMP_FILE_ROOT, file_name)
        if not os.path.exists(abs_filename):
            raise ConnectorError(cs_connector_utility_3.format(abs_filename))
        if file_type == 'XML':
            with open(abs_filename, 'r') as xml_file:
                xml = xml_file.read().replace('\n', '')
                clean_xml = re.sub('\s+', ' ', xml).strip()
                return json.loads(json.dumps(xmltodict.parse(clean_xml)))
        elif file_type == 'CSV':
            with open(abs_filename) as csv_file:
                reader = csv.DictReader(csv_file)
                dict_list = []
                for line in reader:
                    dict_list.append(line)
                return dict_list
        elif file_type == 'XLS, XLSX':
            if file_name.endswith('.xlsx') or abs_filename.endswith('.XLSX'):
                if not os.path.splitext(abs_filename)[1] == '.xlsx':
                    abs_filename_ext = '{}{}'.format(abs_filename, '.xlsx')
                    os.rename(abs_filename, abs_filename_ext)
                    abs_filename = abs_filename_ext

                wb = load_workbook(abs_filename)
                data = OrderedDict()
                for sheet in wb.sheetnames:
                    sheet_data = []
                    title_row = []
                    for row in wb[sheet].values:
                        if not title_row:
                            title_row = list(row)
                            continue
                        row_data = {}
                        row_values = list(row)
                        for cell_index in range(0, len(title_row)):
                            row_data[title_row[cell_index]] = row_values[cell_index]
                        sheet_data.append(row_data)
                    data[sheet] = sheet_data
                return json.loads(json.dumps(data, cls=DjangoJSONEncoder))
            else:
                wb = xlrd.open_workbook(abs_filename)
                data = OrderedDict()
                for sheet_index in range(wb.nsheets):
                    sh = wb.sheet_by_index(sheet_index)
                    if sh.nrows == 0:
                        continue
                    attr_list = sh.row_values(0)
                    rows_list = []
                    for rownum in range(1, sh.nrows):
                        elm = OrderedDict()
                        for index in range(len(attr_list)):
                            elm[attr_list[index]] = sh.cell_value(rownum, index)
                        rows_list.append(elm)
                    data[sh.name] = rows_list
                return json.loads(json.dumps(data))
        else:
            raise ConnectorError(cs_connector_utility_2.format("file type", file_type, "XML/CSV/XLS/XLSL", file_type))
    else:
        raise ConnectorError("%s"%cs_connector_utility_1.format("Either file IRI or file path"))


def html_table_to_dictionary(html, *args, **kwargs):
    """
    Converts a Table from an HTML string into a dictionary

   :param str html: The html containing the table
   :return: The converted dictionary
   :rtype: dict
    """
    # extract table from text
    json_data = []
    rows = BeautifulSoup(html, "lxml")("table")[0]("tr")
    if len(rows[0].find_all("td")):
        keys = [cell.text for cell in rows[0]("td")]
    elif len(rows[0].find_all("th")):
        keys = [cell.text for cell in rows[0]("th")]
    for i in range(1, len(rows)):
        row_dict = {}
        for j, cell in enumerate(rows[i]("td")):
            row_dict[keys[j]] = unescape(cell.text.strip())
        json_data.append(row_dict)
    return json.loads(json.dumps(json_data))


def parse_cef(cef_input, *args, **kwargs):
    """
    Parse a string in CEF format and return a dict with the header values
    and the extension data.
    """

    # Create the empty dict we'll return later
    values = dict()

    # This regex separates the string into the CEF header and the extension
    # data.  Once we do this, it's easier to use other regexes to parse each
    # part.
    header_re = r'(.*(?<!\\)\|){,7}(.*)'

    res = re.search(header_re, cef_input)
    if res:
        header = res.group(1)
        extension = res.group(2)

        # Split the header on the "|" char.  Uses a negative lookbehind
        # assertion to ensure we don't accidentally split on escaped chars,
        # though.
        spl = re.split(r'(?<!\\)\|', header)

        # Since these values are set by their position in the header, it's
        # easy to know which is which.
        values["DeviceVendor"] = spl[1]
        values["DeviceProduct"] = spl[2]
        values["DeviceVersion"] = spl[3]
        values["DeviceEventClassID"] = spl[4]
        values["DeviceName"] = spl[5]
        if len(spl) > 6:
            values["DeviceSeverity"] = spl[6]

        # The first value is actually the CEF version, formatted like
        # "CEF:#".  Ignore anything before that (like a date from a syslog message).
        # We then split on the colon and use the second value as the
        # version number.
        cef_start = spl[0].find('CEF')
        if cef_start == -1:
            return None
        (cef, version) = spl[0][cef_start:].split(':')
        values["CEFVersion"] = version

        # The ugly, gnarly regex here finds a single key=value pair,
        # taking into account multiple whitespaces, escaped '=' and '|'
        # chars.  It returns an iterator of tuples.
        spl = re.findall(r'([^=\s]+)=((?:[\\]=|[^=])+)(?:\s|$)', extension)
        for i in spl:
            # Split the tuples and put them into the dictionary
            values[i[0]] = i[1]

        # Process custom field labels
        result = dict()
        custom_fields = []
        for key in values.keys():
            # If the key string ends with Label, replace it in the appropriate
            #  custom field
            if key.endswith("Label"):
                customfield = key[:-5]
                if customfield in values.keys():
                    custom_fields.append(customfield)
                    result[values[key]] = values[customfield]
            else:
                result[key] = values[key]
        result = {k: v for k, v in result.items() if k not in custom_fields}
    return result


def markdown_to_html(markdown_string=None, *args, **kwargs):
    """
    Converts Markdown string into HTML

   :param str markdown_string: The markdown string to be converted into HTML
   :return: a converted HTML string
   :rtype: string
    """
    if not markdown_string:
        raise ConnectorError('{0}'.format(cs_connector_utility_1.format('Markdown')))
    clean_markdown =  markdown_string.strip()
    return {'html': markdown.markdown(clean_markdown)}